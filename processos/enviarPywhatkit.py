import pywhatkit as kit
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from openpyxl import load_workbook
import re
import sys
#Variáveis 

turma = sys.argv[1] 
mensagem = sys.argv[2] 

def selecionar_arquivo_xlsx(textoDisplay):
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        caminho_arquivo = filedialog.askopenfilename(
            title=textoDisplay,
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )

        if caminho_arquivo is None:
            messagebox.showwarning("Aviso", "Nenhuma entrada foi fornecida.")

        if caminho_arquivo.strip() == "":
            messagebox.showerror("Erro", "O Usuario cancelou o processo.")
            return None
        
        return caminho_arquivo
        
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
        return None
    finally:
        root.destroy()
            
def obter_string_popup(titulo="Entrada", mensagem="Digite algo:"):
    try:
        root = tk.Tk()
        root.withdraw()  # Oculta a janela principal
        root.attributes("-topmost", True)

        resposta = simpledialog.askstring(title=titulo, prompt=mensagem)

        if resposta is None:
            messagebox.showwarning("Aviso", "Nenhuma entrada foi fornecida.")
            return None

        if resposta.strip() == "":
            messagebox.showerror("Erro", "Entrada vazia não é permitida.")
            return None

        return resposta

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
        return None

    finally:
        root.destroy()  # Fecha a janela principal mesmo que ocorra erro

def normalizar(texto):
    if texto is None:
        return None
    else:
        return re.sub(r'\W+', '', str(texto)).upper()

def extrair_turmasPlanilha(caminho_arquivo):
    turmas = []
    contador = 0

    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    for linha in ws.iter_rows(min_row=2, values_only=True):
        codTurma = linha[46]  # 47ª coluna
        codTurma_norm = normalizar(codTurma)

        if any(codTurma_norm in linha for linha in turmas):
            continue
        else:
            turmas.append([contador, codTurma_norm])
            print(f"Turma {codTurma_norm} adicionada.")
            contador += 1

    return turmas

def extrair_matriz_excel(cod_turma, caminho_arquivo):
    total_alunos = 0
    matriz_alunos = []

    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    for linha in ws.iter_rows(2, values_only=True):
        if linha[46] == cod_turma:
            total_alunos += 1
            matricula = linha[4]
            aluno = linha[5]
            telefone1 = linha[33]
            telefone2 = linha[34]
            matriz_alunos.append([total_alunos - 1, matricula, aluno, telefone1, telefone2])
            print(f"Aluno {total_alunos}: {matricula} - {aluno} - {telefone1} - {telefone2}")
        else:
            continue


    return matriz_alunos

caminho_arquivo = selecionar_arquivo_xlsx("Selecione o Relatório de Matrículas Ativas")
if caminho_arquivo == None:
    exit
else:
    turma_normal = normalizar(turma)
    if turma_normal == None:
        messagebox.showerror("Erro",f"Código de Turma inválido")
        exit
    else:
        print(f"Código da turma: {turma_normal}")
        cod_Turmas = extrair_turmasPlanilha(caminho_arquivo)
        if any(turma_normal in linha for linha in cod_Turmas):
            matriz_alunos = extrair_matriz_excel(turma_normal, caminho_arquivo)
            print(f"Total de alunos encontrados: {len(matriz_alunos)}")
            
            pre_delay = 10
            pos_delay = 5
            contador = 0
            for aluno in matriz_alunos:
                nome = aluno[2]
                print(f"Aluno: {nome}")
                mensagem = f"Oi, {nome}"

                # --- Trata número 1 ---
                telefone1 = aluno[3]
                if isinstance(telefone1, str):
                    numero1 = re.sub(r'\D', '', telefone1)
                    if len(numero1) == 11:
                        numero1 = f"+55{numero1}"
                        print(f"numero1: {numero1}")
                    else:
                        numero1 = None
                        print("Número1: inválido")
                else:
                    numero1 = None
                    print("Número1: ausente")

                # --- Trata número 2 ---
                telefone2 = aluno[4]
                if isinstance(telefone2, str):
                    numero2 = re.sub(r'\D', '', telefone2)
                    if len(numero2) == 11:
                        numero2 = f"+55{numero2}"
                        print(f"numero2: {numero2}")
                    else:
                        numero2 = None
                        print("Número2: inválido")
                else:
                    numero2 = None
                    print("Número2: ausente")

                print(f"Enviando mensagem para o aluno {nome}...")

                enviados = set()

                # --- Envia para número 1 ---
                if numero1:
                    try:
                        print(f"Enviando mensagem para {numero1}...")
                        kit.sendwhatmsg_instantly(numero1, mensagem, pre_delay, True, pos_delay)
                        enviados.add(numero1)
                        contador += 1
                    except Exception as e:
                        print(f"Erro ao enviar mensagem para {numero1}: {e}")

                # --- Envia para número 2, se for diferente e válido ---
                if numero2 not in enviados:
                    try:
                        print(f"Enviando mensagem para {numero2}...")
                        kit.sendwhatmsg_instantly(numero2, mensagem, pre_delay, True, pos_delay)
                        contador += 1
                    except Exception as e:
                        print(f"Erro ao enviar mensagem para {numero2}: {e}")

                if not numero1 and not numero2:
                    print(f"Nenhum número válido para o aluno {nome}.")

            print(f"Total de mensagens enviadas: {contador}")
        else:
            messagebox.showerror("Erro",f"Turma {turma} não encontrada")
            exit